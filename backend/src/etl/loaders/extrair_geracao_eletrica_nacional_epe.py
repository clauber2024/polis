"""
EXTRACTOR: Geracao eletrica nacional (EPE/BEN) - denominador da participacao
da MMGD na matriz eletrica (RF-005)
--------------------------------------------------------------------------------
Motivacao: RF-005 (Landing Page) listava "participacao da solar distribuida
na matriz eletrica nacional" em indicadoresIndisponiveis, citando so um
numero estatico (7,0% em 2025) direto no codigo do frontend. Ver
docs/DECISOES.md, ADR "Integracao da participacao da MMGD na matriz eletrica
nacional (EPE/PDGD)".

FONTE NAO E AUTOMATIZAVEL - o Balanco Energetico Nacional (BEN) da EPE nao
tem API REST; o download so existe dentro de um dashboard interativo
(dashboard.epe.gov.br/apps/livro-ben). Este extractor le o arquivo XLSX
baixado manualmente (Anexo X, "Unidades Comerciais", formato tabela/
tidyverse - colunas grupo/fonte/ano/valor), colocado em
backend/src/etl/data/raw/epe_ben_geracao/. Uma atualizacao futura exige
baixar de novo do dashboard e substituir o arquivo.

Linha usada: grupo="Total Transformacao", fonte="Eletricidade - GWh" - e a
geracao eletrica BRUTA total do Brasil (centrais publicas + autoprodutoras,
todas as fontes primarias somadas), nao "Oferta Interna Bruta" (que so mede
saldo de importacao/exportacao/estoque) nem "Consumo Final" (que ja desconta
perdas de distribuicao) - as tres linhas existem na mesma planilha para
"Eletricidade - GWh" e foram inspecionadas nesta sessao antes de escolher
esta. Serie cobre 1970-2025 (56 anos), uma linha por ano, sem furos.

Este extractor carrega SO a coluna geracao_eletrica_nacional_gwh (e a
citacao da fonte). As colunas geracao_mmgd_gwh e
percentual_consumo_cativo_atendido_mmgd (fonte: EPE/PDGD, ver migration 0030)
sao carregadas por um extractor separado, ainda nao escrito (arquivo do PDGD
pendente) - o ON CONFLICT abaixo NAO toca nessas colunas, para nao apagar um
valor ja carregado por aquele outro extractor.
"""
import os
import sys

import openpyxl
from sqlalchemy import create_engine, text

DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql://atlas:atlas_dev_local@localhost:5432/atlas_solar_justo"
)
CAMINHO_XLSX = os.path.join(
    os.path.dirname(__file__),
    "..", "data", "raw", "epe_ben_geracao", "tabela_balanco_energitico_unidades_comerciais.xlsx",
)
FONTE_CITACAO = (
    "EPE, Balanco Energetico Nacional (BEN), Anexo X - Unidades Comerciais, "
    "grupo=Total Transformacao, fonte=Eletricidade - GWh"
)
GRUPO_ALVO = "Total Transformação"
FONTE_ALVO_PREFIXO = "Eletricidade"


def ler_serie_geracao_nacional(caminho: str) -> dict[str, float]:
    """Le a planilha (formato longo grupo/fonte/ano/valor) e retorna
    {ano: geracao_gwh} so para grupo=Total Transformacao e fonte iniciando
    com 'Eletricidade' - ver docstring do modulo para o porque desta linha."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    cabecalho = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    colunas_esperadas = ("grupo", "fonte", "ano", "valor")
    if tuple(cabecalho) != colunas_esperadas:
        raise ValueError(
            f"Cabecalho inesperado: {cabecalho!r} (esperado {colunas_esperadas!r}). "
            "O formato do export do BEN pode ter mudado - conferir manualmente antes de prosseguir."
        )

    serie: dict[str, float] = {}
    for grupo, fonte, ano, valor in ws.iter_rows(min_row=2, values_only=True):
        if grupo == GRUPO_ALVO and fonte and str(fonte).startswith(FONTE_ALVO_PREFIXO):
            serie[str(ano)] = float(valor)
    return serie


def main():
    print("Extractor Geracao Eletrica Nacional (EPE/BEN)")
    print("=" * 70)

    if not os.path.exists(CAMINHO_XLSX):
        print(f"[ERRO] Arquivo nao encontrado: {CAMINHO_XLSX}")
        print("       Baixe o Anexo X (Unidades Comerciais, formato tabela/tidyverse) em")
        print("       dashboard.epe.gov.br/apps/livro-ben/livro/pt/anexo_10.html e salve nesse caminho.")
        sys.exit(1)

    print(f"[1/2] Lendo {CAMINHO_XLSX} ...")
    serie = ler_serie_geracao_nacional(CAMINHO_XLSX)
    if not serie:
        print(f"[ERRO] Nenhuma linha encontrada para grupo={GRUPO_ALVO!r} / fonte iniciando com {FONTE_ALVO_PREFIXO!r}.")
        sys.exit(1)
    anos_ordenados = sorted(serie)
    print(f"      {len(serie)} anos encontrados ({anos_ordenados[0]}-{anos_ordenados[-1]}).")
    print(f"      Ultimo ano: {anos_ordenados[-1]} = {serie[anos_ordenados[-1]]:,.1f} GWh")

    engine = create_engine(DATABASE_URL)
    sql_upsert = text("""
        INSERT INTO indicadores_energia_nacional
            (periodo_referencia, geracao_eletrica_nacional_gwh, fonte_geracao_nacional)
        VALUES
            (:periodo, :valor, :fonte)
        ON CONFLICT (periodo_referencia) DO UPDATE SET
            geracao_eletrica_nacional_gwh = EXCLUDED.geracao_eletrica_nacional_gwh,
            fonte_geracao_nacional = EXCLUDED.fonte_geracao_nacional,
            atualizado_em = now()
    """)

    print(f"[2/2] Inserindo/atualizando {len(serie)} anos...")
    inseridos = 0
    falhas = []
    for ano in anos_ordenados:
        try:
            with engine.begin() as con:
                con.execute(
                    sql_upsert,
                    {
                        "periodo": f"{ano}-01-01",
                        "valor": round(serie[ano], 3),
                        "fonte": FONTE_CITACAO,
                    },
                )
            inseridos += 1
        except Exception as e:
            falhas.append((ano, str(e)[:120]))

    print(f"      {inseridos} anos inseridos/atualizados. Falhas: {len(falhas)}")
    for ano, erro in falhas[:5]:
        print(f"        - {ano}: {erro}")

    print("Extractor Geracao Eletrica Nacional (EPE/BEN) concluido.")


if __name__ == "__main__":
    main()
