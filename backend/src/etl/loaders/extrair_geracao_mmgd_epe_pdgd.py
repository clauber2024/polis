"""
EXTRACTOR: Geracao estimada de MMGD (EPE/PDGD) - numerador da participacao
da MMGD na matriz eletrica nacional (RF-005)
--------------------------------------------------------------------------------
Motivacao: complementa extrair_geracao_eletrica_nacional_epe.py (denominador,
BEN). Ver docs/DECISOES.md, ADR "Integracao da participacao da MMGD na
matriz eletrica nacional (EPE/PDGD)".

FONTE NAO E AUTOMATIZAVEL - o Painel de Dados de MMGD (PDGD) da EPE e uma
app Shiny sem API/URL de download estavel (dashboard.epe.gov.br/apps/pdgd,
aba "Geracao de Eletricidade"). Este extractor le o arquivo XLSX baixado
manualmente pelo botao "Baixar Dados dos Graficos" do grafico "Estimativa
da Geracao no Ano" (dados desagregados por subsistema/UF/classe/fonte/
modalidade/subgrupo/distribuidora/segmento/mini_micro/ano_operacao/
autoc_inj), colocado em
backend/src/etl/data/raw/epe_pdgd_geracao/dados.xlsx. Uma atualizacao futura
exige baixar de novo do dashboard e substituir o arquivo.

A geracao TOTAL de MMGD por ano = soma de energia_gwh_div de TODAS as linhas
daquele ano_operacao, somando as duas categorias de autoc_inj
(autoconsumo_gwh + injecao_gwh - decomposicao mutuamente exclusiva da mesma
geracao, nao duplicata). Nao filtra por UF/classe/fonte/modalidade - soma
tudo para chegar no total nacional. Conferido nesta sessao: o total de 2025
(~54.483 GWh) bate visualmente com a barra do grafico "Estimativa da
Geracao no Ano" no dashboard.

Este extractor carrega SO geracao_mmgd_gwh (e a citacao da fonte) - NAO
carrega percentual_consumo_cativo_atendido_mmgd (grafico "% do Consumo
Cativo BR", outro botao de download dentro da mesma aba, arquivo ainda nao
obtido - ver docs/PLANO_ATUAL.md). O ON CONFLICT abaixo NAO toca em
geracao_eletrica_nacional_gwh/fonte_geracao_nacional (carregados por
extrair_geracao_eletrica_nacional_epe.py) nem em
percentual_consumo_cativo_atendido_mmgd/fonte_mmgd quando esta ja tiver sido
preenchida por outro extractor no futuro - so fonte_mmgd e reescrita aqui
porque e a mesma citacao (PDGD) usada por ambas as colunas de MMGD.

IMPORTANTE: a serie so cobre 2013-2025 (ano em que a MMGD comecou a existir
no Brasil) - anos anteriores (presentes na tabela do BEN) ficam sem
geracao_mmgd_gwh carregado por este extractor (permanecem NULL, nao zero -
MMGD simplesmente nao existia, mas tambem nao fazia parte do universo desta
tabela antes do BEN ja ter inserido a linha do ano via o outro extractor).
"""
import os
import sys
from collections import defaultdict

import openpyxl
from sqlalchemy import create_engine, text

DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql://atlas:atlas_dev_local@localhost:5432/atlas_solar_justo"
)
CAMINHO_XLSX = os.path.join(
    os.path.dirname(__file__),
    "..", "data", "raw", "epe_pdgd_geracao", "dados.xlsx",
)
FONTE_CITACAO = (
    "EPE, Painel de Dados de MMGD (PDGD), aba Geracao de Eletricidade, "
    "grafico Estimativa da Geracao no Ano (soma nacional de autoconsumo + injecao na rede)"
)
COLUNAS_ESPERADAS = (
    "subsistema", "uf", "classe", "fonte_resumo", "modalidade", "subgrupo",
    "distribuidora", "nome_4md", "ano_operacao", "segmento", "mini_micro",
    "potencia_mw", "autoc_inj", "energia_gwh_div", "share_dist", "share_classe", "share_br",
)


def ler_serie_geracao_mmgd(caminho: str) -> dict[int, float]:
    """Le a planilha desagregada e retorna {ano: geracao_mmgd_gwh} somando
    energia_gwh_div de todas as linhas por ano_operacao - ver docstring do
    modulo para o porque de somar tudo (nenhum filtro reduz para o total
    nacional)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    cabecalho = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if tuple(cabecalho) != COLUNAS_ESPERADAS:
        raise ValueError(
            f"Cabecalho inesperado: {cabecalho!r} (esperado {COLUNAS_ESPERADAS!r}). "
            "O formato do export do PDGD pode ter mudado - conferir manualmente antes de prosseguir."
        )
    idx_ano = COLUNAS_ESPERADAS.index("ano_operacao")
    idx_valor = COLUNAS_ESPERADAS.index("energia_gwh_div")

    por_ano: dict[int, float] = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ano = row[idx_ano]
        valor = row[idx_valor]
        if ano is not None and valor:
            por_ano[int(ano)] += float(valor)
    return dict(por_ano)


def main():
    print("Extractor Geracao Estimada de MMGD (EPE/PDGD)")
    print("=" * 70)

    if not os.path.exists(CAMINHO_XLSX):
        print(f"[ERRO] Arquivo nao encontrado: {CAMINHO_XLSX}")
        print("       Baixe em dashboard.epe.gov.br/apps/pdgd, aba 'Geracao de Eletricidade',")
        print("       botao 'Baixar Dados dos Graficos' do grafico 'Estimativa da Geracao no Ano',")
        print("       e salve nesse caminho com o nome dados.xlsx.")
        sys.exit(1)

    print(f"[1/2] Lendo {CAMINHO_XLSX} ...")
    serie = ler_serie_geracao_mmgd(CAMINHO_XLSX)
    if not serie:
        print("[ERRO] Nenhuma linha com energia_gwh_div valida encontrada.")
        sys.exit(1)
    anos_ordenados = sorted(serie)
    print(f"      {len(serie)} anos encontrados ({anos_ordenados[0]}-{anos_ordenados[-1]}).")
    print(f"      Ultimo ano: {anos_ordenados[-1]} = {serie[anos_ordenados[-1]]:,.1f} GWh")

    engine = create_engine(DATABASE_URL)
    sql_upsert = text("""
        INSERT INTO indicadores_energia_nacional
            (periodo_referencia, geracao_mmgd_gwh, fonte_mmgd)
        VALUES
            (:periodo, :valor, :fonte)
        ON CONFLICT (periodo_referencia) DO UPDATE SET
            geracao_mmgd_gwh = EXCLUDED.geracao_mmgd_gwh,
            fonte_mmgd = EXCLUDED.fonte_mmgd,
            atualizado_em = now()
    """)

    print(f"[2/2] Inserindo/atualizando {len(serie)} anos...")
    inseridos = 0
    falhas = []
    for ano in anos_ordenados:
        try:
            with engine.begin() as con:
                con.execute(
                    sql_upsert,
                    {
                        "periodo": f"{ano}-01-01",
                        "valor": round(serie[ano], 3),
                        "fonte": FONTE_CITACAO,
                    },
                )
            inseridos += 1
        except Exception as e:
            falhas.append((ano, str(e)[:120]))

    print(f"      {inseridos} anos inseridos/atualizados. Falhas: {len(falhas)}")
    for ano, erro in falhas[:5]:
        print(f"        - {ano}: {erro}")

    print("Extractor Geracao Estimada de MMGD (EPE/PDGD) concluido.")


if __name__ == "__main__":
    main()
